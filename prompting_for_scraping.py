from openai import OpenAI
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import os
import json


load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")

client = OpenAI(
    api_key=api_key,
    base_url="https://api.sambanova.ai/v1"
)


prompt = """
You are a data extraction system.

Your task is to extract structured JSON from web-scraped company data. Follow all rules strictly. Output ONLY a valid JSON object. 

Do NOT include:
- Explanations
- Code
- Markdown (```)
- Comments
- Anything except raw JSON

RULES FOR OUTPUT:
- Output only valid JSON in the structure below
- Use only facts explicitly stated in the data
- Do NOT guess or fill in information
- If a value is missing, use:
  - "" for text fields
  - [] for lists
  - null for unit_number

VALID CATEGORIES:
["roofing", "flooring", "electrical work", "decks & balconies", "transportation", "windows", "handyman", "painting & wall finishes", "foundation", "doors", "masonry", "drainage", "smart home", "basement", "demolition", "excavation", "plumbing", "commercial cleaning", "garage", "energy advisors", "kitchen", "residential cleaning", "snow removal", "bathroom", "landscaping", "hvac", "general contractor", "siding"]

BIO RULES:
- Must be 500–1000 characters
- Use third-person, professional tone
- Use only information present in the input
- Focus on services, experience, and location

SOCIAL LINKS:
Extract URLs if they contain:
- facebook.com or fb.com
- instagram.com
- homestars.com
- google.com/maps or maps.google.com
- twitter.com or x.com

LOCATIONS:
- Canada addresses only
- Format postal codes as "A1A 1A1"
- Remove duplicates

LANGUAGES:
- List all languages clearly visible in the data
- Include "English" if any English content exists

OUTPUT FORMAT (must match exactly):

{
  "company_name": "",
  "website_url": "",
  "company_logo": "",
  "languages": [],
  "categories": [],
  "company_bio": "",
  "locations": [
    {
      "street_address": "",
      "unit_number": null,
      "postal_code": "",
      "country": "Canada",
      "province": "",
      "city": "",
      "contact_name": "",
      "contact_phone_number": "",
      "property_notes": ""
    }
  ],
  "social_media": {
    "facebook": "",
    "instagram": "",
    "homestars": "",
    "google": "",
    "twitter": ""
  }
}

Now begin.

Web-scraped data:

"""


def generate_json_from_data(data, logos):

    company_name = client.chat.completions.create(
            model="Meta-Llama-3.1-8B-Instruct",
            messages=[
                {
                    "role": "user",
                    "content": data + " from the above data, find out what is most likely the company name and return it, do not return any other information or formating, just the company name"
                }
            ]
        )
    
    company_name = company_name.choices[0].message.content.strip()
    
    logo = find_right_logo(logos, company_name)

    response = client.chat.completions.create(
            model="Meta-Llama-3.1-8B-Instruct",
            messages=[
                {
                    "role": "user",
                    "content": prompt + data + " below is the logo for the company " + logo
                }
            ]
        )
    

    return response.choices[0].message.content.strip()


prompt2 = """You will be given a list of logo URLs and a company name. Your task is to determine which logo best represents the company based on the provided URLs and find the most appropriate logo. Output only the URL of the selected logo, with no extra text, explanation, or formatting. you must pick 1 of the logos fromt the list which will be given to you The company name is: """

def find_right_logo(logos, company_name):

    logo_options = ''.join(map(str, logos))

    response = client.chat.completions.create(
        model="Meta-Llama-3.1-8B-Instruct",
        messages=[
            {
                "role": "user",
                "content": prompt2 + company_name +  logo_options
            }
        ]
    )

    return response.choices[0].message.content.strip()


get_service_prompt = "You will be given data from a website, your job is to go through the data and find out what services the company provides, then to generate an 50 - 80 charecter description for each service, you should only return the list in the following format with no internal throughts or extra formating: this is the format service_1: description| service_2: description| service_3: description here is an examplar of the format you should return: roofing: We provide high-quality roofing services including installation, repair, and maintenance. | flooring: Expert flooring solutions for residential and commercial properties. now here is the data: "


def estimate_services_format(data):
    response = client.chat.completions.create(
        model="Meta-Llama-3.1-8B-Instruct",
        messages=[
            {
                "role": "user",
                "content":  get_service_prompt + data
            }
        ]
    )

    string_services = response.choices[0].message.content.strip()
    raw_services = string_services.split('|')

    services = []
    for service in raw_services:
        if ':' in service:
            name, description = service.split(':', 1)
            services.append({
                'name': name.strip(),
                'description': description.strip()
            })

    make_excel(services)

    return services

def make_excel(items):
    output_filename="services.xlsx"

    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = 'Service Name'
    ws['B1'] = 'Service Description'
    ws['C1'] = 'Service Price'

    bold_font = Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font
    ws['C1'].font = bold_font

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 20
        
    for i in range(len(items)):
        row_num = i + 2  
        
        ws[f'A{row_num}'] = items[i]['name']        
        ws[f'B{row_num}'] = items[i]['description']
        ws[f'C{row_num}'] = ""
    
        ws[f'B{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_filename)

    return output_filename



def make_json(output, company_name):
    try:
        parsed_json = json.loads(output)
        required_fields = ["categories"]
        all_required_present = True
        
        for field in required_fields:
            value = parsed_json.get(field)
            if not value or value == "null":
                all_required_present = False
                break
        
        if all_required_present:
            with open( f"{company_name}.json", "w") as f:
                json.dump(parsed_json, f, indent=4)
                    
    except json.JSONDecodeError:
        pass